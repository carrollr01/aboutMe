#!/usr/bin/env python3
"""Build the weekly deals workbook: M&A tab + Growth Capital Raises tab, links shown as 'Link'."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MA_HEADERS = ["Sector", "Target Country", "Deal Type", "Date", "Target", "Acquirer",
              "EV ($M)", "EV / Revenue", "EV / EBITDA", "Target Description", "Link / Press Release"]
RAISE_HEADERS = ["Sector", "Target Country", "Date", "Target", "Lead Investor(s)",
                 "Amount ($M)", "Valuation ($M)", "EV / Revenue", "EV / EBITDA",
                 "Target Description", "Link / Press Release"]

# M&A rows: [Sector, Country, DealType, Date, Target, Acquirer, EV($M), EV/Rev, EV/EBITDA, Desc, URL]
MA = [
 ["Asset & Wealth Tech","Switzerland","Strategic M&A","08-Jun-26","additiv","Temenos","","","",
  "Provider of an AI-driven orchestration platform that lets banks, insurers, and wealth managers launch and run digital wealth, banking, and investment products",
  "https://www.temenos.com/press_release/temenos-acquires-additiv-to-strengthen-its-wealth-proposition-and-accelerate-ai-driven-orchestration/"],
 ["Capital Markets Tech","United States","Strategic M&A","08-Jun-26","LevPro","Octus","","","",
  "Provider of front-office portfolio management and trading software purpose-built for CLO and private credit managers",
  "https://octus.com/company/news/octus-signs-definitive-agreement-to-acquire-levpro/"],
 ["Corporate Financial Function","United States","Strategic M&A","11-Jun-26","Orb","Adyen",335,"","",
  "Enterprise billing and revenue-management platform that turns complex usage- and outcome-based pricing into accurate invoices for software companies",
  "https://www.adyen.com/press-and-media/jtrg4qd7j3p4rj"],
 ["Corporate Financial Function","United Kingdom","Strategic M&A","08-Jun-26","m3ter","Salesforce","","","",
  "Usage-based billing and metering platform that powers consumption- and AI-based pricing models for enterprise software companies",
  "https://www.salesforce.com/news/stories/salesforce-signs-definitive-agreement-to-acquire-m3ter/"],
 ["Financial Info & Analytics","United States","Strategic M&A","11-Jun-26","Messari","Blockworks",10,"","",
  "Crypto market-intelligence and data platform covering 40,000+ digital assets through markets, indices, research, and a data API",
  "https://www.businesswire.com/news/home/20260611031601/en/Blockworks-Acquires-Messari-Combining-the-Two-Largest-Crypto-Data-Platforms"],
 ["Financial Info & Analytics","United States","Strategic M&A","10-Jun-26","RiskFront AI","K2 Integrity","","","",
  "Agentic AI risk solution designed to help banks detect and prevent financial crime and automate AML compliance without adding friction for genuine customers",
  "https://fintech.global/2026/06/10/k2-integrity-targets-financial-crime-with-ai-deal/"],
 ["Payments","United States","Strategic M&A","12-Jun-26","Payoneer","Nuvei",2750,"","",
  "Cross-border payments and money-movement platform serving small businesses and online marketplace sellers across more than 190 countries",
  "https://www.prnewswire.com/news-releases/nuvei-to-acquire-payoneer-for-2-75-billion-creating-a-leading-global-platform-for-local-and-cross-border-commerce-302800166.html"],
 ["Real Estate & Mortgage Tech","United States","Strategic M&A","10-Jun-26","Kiavi","Figure",717,"","",
  "AI-powered non-bank lending platform providing fix-and-flip and rental financing to residential real-estate investors",
  "https://www.kiavi.com/press/figure-acquires-kiavi"],
]

# Raise rows: [Sector, Country, Date, Target, Lead, Amount($M), Valuation($M), EV/Rev, EV/EBITDA, Desc, URL]
RAISES = [
 ["Asset & Wealth Tech","United States","11-Jun-26","Hypha","TriEdge Investments",50,"","","",
  "AI-native fund operations platform targeting private markets, handling documentation and data with human oversight",
  "https://www.businesswire.com/news/home/20260611628926/en/Hypha-Emerges-From-Stealth-Announces-a-$50M-Seed-Round"],
 ["Banking & Lending Tech","United States","12-Jun-26","Current","Springcoast Partners",80,1500,"","",
  "Digital banking provider offering spending, savings, and credit-building accounts tailored to everyday US consumers",
  "https://www.thebankslate.com/2026/06/neobank-current-raises-80-million-in-latest-funding-round/"],
 ["Banking & Lending Tech","United States","08-Jun-26","EDGE Markets","CoinFund",29.2,"","","",
  "Builder of banking and payment-rail infrastructure (EDGE Connect and EDGE Pro) for CFTC-regulated prediction markets and gaming",
  "https://www.prnewswire.com/news-releases/edge-markets-raises-29-2-million-series-a-funding-round-302793671.html"],
 ["Capital Markets Tech","United States","11-Jun-26","Digital Asset","a16z crypto",355,2000,"","",
  "Developer of the Canton Network, a public Layer-1 blockchain bringing institutional capital markets onchain",
  "https://www.prnewswire.com/news-releases/digital-asset-raises-355-million-to-accelerate-cantons-role-as-onchain-infrastructure-for-capital-markets-302797427.html"],
 ["Financial Info & Analytics","United States","09-Jun-26","NinjaOne","Wellington Management, ICONIQ, Sequoia",400,12300,"","",
  "Unified IT operations and endpoint-management platform serving nearly 40,000 organizations across 140+ countries",
  "https://www.ninjaone.com/press/12-3-billion-valuation/"],
 ["InsurTech","United States","10-Jun-26","Poetic","Kleiner Perkins",50,500,"","",
  "Developer of deterministic AI that automates complex, multi-step insurance and enterprise back-office processes",
  "https://www.prnewswire.com/news-releases/poetic-raises-50m-series-a-to-automate-the-worlds-most-complex-enterprise-processes-with-reliable-ai-302796939.html"],
 ["Payments","Egypt","08-Jun-26","Blnk","Algebra Ventures",37,"","","",
  "AI-powered point-of-sale and consumer financing platform extending installment credit through thousands of merchants across Egypt",
  "https://ffnews.com/newsarticle/funding/blnk-raises-37m-to-deepen-point-of-sale-credit-for-consumers-in-egypt/"],
 ["Real Estate & Mortgage Tech","Canada","10-Jun-26","nesto","La Caisse",220,1070,"","",
  "AI-driven mortgage technology and financing platform combining direct lending with white-label cloud software for lenders",
  "https://www.globenewswire.com/news-release/2026/06/10/3309762/0/en/nesto-raises-302-million-Series-E-at-1-47-billion-valuation-to-accelerate-growth.html"],
]

header_fill = PatternFill("solid", fgColor="1F6FB2")
header_font = Font(color="FFFFFF", bold=True, size=11)
link_font = Font(color="0563C1", underline="single", size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill_sheet(ws, headers, rows, widths):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center; cell.border = border
    link_col = len(headers)  # last column is the link
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = border
            cell.alignment = left_wrap
            if ci == link_col:
                cell.value = "Link"
                cell.hyperlink = val
                cell.font = link_font
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.value = val
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{1 + len(rows)}"
    ws.row_dimensions[1].height = 30


wb = Workbook()
ws_ma = wb.active
ws_ma.title = "M&A"
fill_sheet(ws_ma, MA_HEADERS, MA, [24, 15, 15, 11, 16, 18, 9, 11, 11, 62, 10])
ws_r = wb.create_sheet("Growth Capital Raises")
fill_sheet(ws_r, RAISE_HEADERS, RAISES, [24, 15, 11, 16, 26, 11, 13, 11, 11, 62, 10])

out = "/home/user/aboutMe/Weekly_Fintech_Deals_formatted.xlsx"
wb.save(out)
print("Wrote", out, "-", len(MA), "M&A +", len(RAISES), "raises")
