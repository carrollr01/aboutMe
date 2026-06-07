#!/usr/bin/env python3
"""Build the verified fintech deals Excel sheet."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADERS = ["Sector", "Target Country", "Deal Type", "Date", "Target", "Acquirer",
           "EV ($M)", "EV / Revenue", "EV / EBITDA", "Target Description",
           "Link / Press Release"]

# Each row uses the VERIFIED (corrected) date. EV ($M) filled for M&A where a deal
# value is known; blank for capital raises (raise size noted in description).
# EV/Revenue and EV/EBITDA left blank — not disclosed for any of these deals.
ROWS = [
 ["Asset & Wealth Tech","United States","Strategic M&A","2026-06-03","Affinity Advisory Network / AAN Wealth Advisors","Nu Ride Inc.",9.6,"","",
  "Integrated insurance distribution + RIA advisory network serving agents/advisors across 700+ US cities; deal ~$9.6M incl. cash, stock and earnout.",
  "https://www.investing.com/news/company-news/nu-ride-to-acquire-affinity-advisory-for-96-million-93CH-4724416"],
 ["Banking & Lending Tech","Australia","Strategic M&A","2026-06-05","Interfi (Interfi Systems)","Perpetual","","","",
  "Loan-servicing technology for non-bank lenders; ~A$55bn AUA across the full loan lifecycle. Perpetual took an initial 70% stake.",
  "https://www.financialstandard.com.au/news/perpetual-acquires-majority-stake-in-interfi-179812809"],
 ["Banking & Lending Tech","United States","Strategic M&A","2026-06-03","Finastra US Mid-Market Banking business","CORA Group (Jonas Software / Constellation)","","","",
  "Core & digital banking software (Phoenix Core System, Malauzai Digital Banking, Fusion Analytics) used by hundreds of US banks/credit unions.",
  "https://www.globenewswire.com/news-release/2026/06/03/3306031/0/en/cora-group-acquires-finastra-s-phoenix-core-system-malauzai-digital-banking-and-fusion-analytics-businesses.html"],
 ["Banking & Lending Tech","Kenya","Strategic M&A","2026-03-31","Sumac Microfinance Bank","Moniepoint","","","",
  "Kenyan microfinance bank (78% stake acquired) marking Moniepoint's entry into East Africa. NOTE: announced/closed late MARCH 2026, not June.",
  "https://www.businesswire.com/news/home/20260331855947/en/Moniepoint-Inc.-Enters-Kenyan-Market-With-Acquisition-of-Sumac-Microfinance-Bank"],
 ["Banking & Lending Tech","United States","Capital Raise","2026-05-27","Capchase","01 Advisors (lead)","","","",
  "B2B / vendor financing platform ('Affirm for B2B'). Raised $26M equity + $174M credit facility.",
  "https://news.crunchbase.com/venture/fintech-capchase-b2b-bnpl-200m-debt-equity/"],
 ["Banking & Lending Tech","India","Capital Raise","2026-06-04","WeRize","Sony Innovation Fund (lead)","","","",
  "AI-powered credit, insurance and investment platform for small-city ('Bharat') India. Raised $7M pre-Series C.",
  "https://inc42.com/buzz/fintech-startup-werize-raises-7-mn-to-expand-product-offerings/"],
 ["Capital Markets Tech","Netherlands","Capital Raise","2026-06-04","Cense","G+D Ventures & Rabo Investments (co-lead)","","","",
  "Digital-asset compliance & evidence platform for banks/financial institutions. Raised EUR 6.5M seed. (Closer to RegTech / Financial Info than Capital Markets.)",
  "https://fintech.global/2026/06/04/cense-raises-e6-5m-seed-round-for-digital-asset-compliance/"],
 ["Capital Markets Tech","United Kingdom","Strategic M&A","2026-05-26","Funded Trading Plus","Instant Funding","","","",
  "Proprietary trading firm offering funded-trader evaluation products and funded trading accounts.",
  "https://windsordrake.com/market-intelligence/transactions/funded-trading-plus-instant-funding-2026"],
 ["Corporate Financial Function","United States","Capital Raise","2026-06-04","Ramp","ICONIQ, GIC, Ontario Teachers' (lead)","","","",
  "AI spend-management / financial-operations platform (corporate cards, expense, AP, treasury). Raised $750M Series F at $44B valuation.",
  "https://www.prnewswire.com/news-releases/ramp-raises-series-f-at-44-billion-valuation-302791103.html"],
 ["Corporate Financial Function","United States","Strategic M&A","2026-06-04","Leapfin","Airwallex","","","",
  "Revenue-recognition & reconciliation automation platform (record-to-report; GAAP/IFRS-ready financials).",
  "https://www.airwallex.com/global/newsroom/airwallex-acquires-leapfin-expanding-financial-lifecycle-capabilities"],
 ["Corporate Financial Function","France","Capital Raise","2026-05-21","Pivot","Forestay Capital & Notion Capital (lead)","","","",
  "AI enterprise procurement / spend operating system (Paris-based). Raised $40M Series B. NOTE: source data listed US — company is French.",
  "https://www.globenewswire.com/news-release/2026/05/21/3299278/0/en/Pivot-Raises-40-Million-Series-B-to-Replace-Legacy-Procurement-Software-with-an-Enterprise-AI-Operating-System.html"],
 ["Corporate Financial Function","Belgium","Capital Raise","2026-05-22","Harmoney","Smile Sail","","","",
  "KYC/AML/counterparty-risk compliance orchestration platform (Ghent). EUR 10M strategic minority investment. (RegTech.)",
  "https://fintech.global/2026/05/26/harmoney-eyes-european-growth-with-e10m-investment/"],
 ["Financial Info & Analytics","United States","Capital Raise","2026-06-03","AlphaSense","Vitruvian Partners, Accenture Ventures, J.P. Morgan Asset Mgmt (lead)","","","",
  "AI market-intelligence / research & workflow platform. Raised $350M at $7.5B valuation; >$600M ARR.",
  "https://www.globenewswire.com/news-release/2026/06/03/3305968/0/en/alphasense-raises-350m-at-7-5b-valuation-and-surpasses-600m-in-annual-recurring-revenue.html"],
 ["Financial Info & Analytics","United States","Strategic M&A","2026-06-04","PEER DATA","Lukka","","","",
  "Data provenance / Data Book of Record (DBOR) and compliance platform for institutional digital assets.",
  "http://www.prnewswire.com/news-releases/lukka-acquires-peer-data-to-build-the-institutional-control-layer-for-digital-assets-and-data-commerce-302791630.html"],
 ["Financial Info & Analytics","United Kingdom","Capital Raise","2026-06-04","Aveni","PXN Ventures (lead)","","","",
  "AI compliance / assurance platform for financial services (Edinburgh). Raised GBP 12M.",
  "https://fintech.global/2026/06/04/lloyds-and-nationwide-backed-aveni-raises-12m/"],
 ["InsurTech","United States","Capital Raise","2026-06-04","Honeycomb Insurance","Zeev Ventures (lead)","","","",
  "AI underwriting platform for commercial property / real-estate (multifamily) insurance. Raised $40M.",
  "https://fortune.com/2026/06/04/honeycomb-insurance-ai-apartments-40-million/"],
 ["InsurTech","United States","Capital Raise","2026-05-28","Corgi","TCV (lead)","","","",
  "Full-stack commercial insurance platform. Raised $106M Series B1 at $2.6B valuation (3 weeks after a $160M Series B).",
  "https://techcrunch.com/2026/05/28/corgi-announces-106m-raise-at-2-6b-valuation-three-weeks-after-160m-series-b/"],
 ["InsurTech","United States","Strategic M&A","2026-06-01","Aggne (Aggne Global)","Wipro",28.5,"","",
  "P&C insurance technology / consulting services. Wipro raised its stake from 60% to 80% for $28.5M.",
  "https://www.business-standard.com/amp/companies/news/wipro-raises-stake-in-insurance-tech-firm-aggne-to-80-for-28-5-million-126060102064_1.html"],
 ["Payments","Canada","Strategic M&A","2026-03-30","KUBRA (Kubra Data Transfer)","REPAY",372,"","",
  "Customer-experience billing & payments technology for utilities/government. Announced 30 Mar 2026 (closed 1 Jun). Mississauga, Canada-based.",
  "https://investors.repay.com/news-releases/news-release-details/repay-announces-agreement-acquire-kubra"],
 ["Payments","United States","Strategic M&A","2026-06-02","Fee Navigator","NMI","","","",
  "AI merchant-pricing intelligence (analyzes merchant statements to recommend optimized pricing).",
  "https://www.businesswire.com/news/home/20260602447713/en/NMI-Acquires-Fee-Navigator-Adding-AI-Powered-Pricing-Intelligence-to-Its-Embedded-Payments-Platform"],
 ["Payments","United Kingdom","SPAC merger (de-SPAC)","2026-06-01","OpenPayd","Titan Acquisition Corp (SPAC)",1100,"","",
  "Embedded finance / fiat & stablecoin payment orchestration. Nasdaq listing via SPAC at ~$1.1B equity value (NOT a strategic acquisition).",
  "https://www.globenewswire.com/news-release/2026/06/01/3304275/0/en/OpenPayd-Targets-Nasdaq-Listing-at-Unicorn-Valuation.html"],
 ["Payments","Canada","Strategic M&A (closed; announced 2025)","2026-06-01","WonderFi","Robinhood",182,"","",
  "Canadian regulated crypto exchanges (Bitbuy, Coinsquare). ~C$250M (~US$182M). NOTE: announced 2025; 1 Jun = CLOSE, not announcement. Crypto exchange (arguably Capital Markets).",
  "https://www.newsfilecorp.com/release/299681/Robinhood-Completes-Acquisition-of-WonderFi"],
 ["Payments","United Kingdom","Strategic M&A","2026-06-02","Absolute Payment Solutions","XFolio AI","","","",
  "Pay.UK-accredited Bacs service provider; combined with XFolio AI (Paris) treasury & payments platform.",
  "https://www.fintechfutures.com/m-a/xfolio-ai-acquires-absolute-payments-solutions"],
 ["Real Estate & Mortgage Tech","Canada","Strategic M&A","2026-05-20","Accepted Financial Corp","Shelter Lending Corporation","","","",
  "Western-Canada alternative mortgage lender/broker (Langley, BC). Services business rather than pure tech.",
  "https://www.cantechletter.com/newswires/shelter-lending-corporation-announces-acquisition-of-accepted-financial-corp-accelerating-national-growth-strategy/"],
]

wb = Workbook()
ws = wb.active
ws.title = "Verified Deals Jun 2026"

header_fill = PatternFill("solid", fgColor="1F6FB2")
header_font = Font(color="FFFFFF", bold=True, size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(HEADERS)
for c, _ in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=c)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border

for r in ROWS:
    ws.append(r)

# style body
for row in ws.iter_rows(min_row=2, max_row=1+len(ROWS), min_col=1, max_col=len(HEADERS)):
    for cell in row:
        cell.alignment = left_wrap
        cell.border = border

widths = [22, 16, 22, 12, 30, 32, 9, 12, 12, 60, 48]
from openpyxl.utils import get_column_letter
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{1+len(ROWS)}"
ws.row_dimensions[1].height = 30

wb.save("/home/user/aboutMe/fintech-deals-verified.xlsx")
print("Wrote fintech-deals-verified.xlsx with", len(ROWS), "rows")
