#!/usr/bin/env python3
"""Research valuation multiples for the precedent deals in deals.json.

For each deal, runs a Claude web-search research pass that determines whether
a revenue multiple or EV multiple was published for the transaction, or can be
calculated from published information (deal value + disclosed target
financials), and records sources.

Usage:
    python3 research_multiples.py                  # research all pending deals
    python3 research_multiples.py --deal nasdaq-adenza
    python3 research_multiples.py --limit 5        # first 5 pending deals
    python3 research_multiples.py --force          # re-run even if a result exists
    python3 research_multiples.py --dry-run        # show the plan, no API calls
    python3 research_multiples.py --summarize-only # rebuild CSV/MD from saved results

Requires: pip install anthropic, and ANTHROPIC_API_KEY in the environment
(or an `ant auth login` profile). Each deal is capped at 3 web searches and
runs at medium effort, so expect roughly $0.10-0.30 per deal; results are
saved per-deal under outputs/results/ so an interrupted run resumes where it
left off.
"""

import argparse
import csv
import json
import re
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
DEALS_FILE = HERE / "deals.json"
RESULTS_DIR = HERE / "outputs" / "results"
SUMMARY_CSV = HERE / "outputs" / "deal_multiples_summary.csv"
SUMMARY_MD = HERE / "outputs" / "deal_multiples_summary.md"
SUMMARY_XLSX = HERE / "outputs" / "deal_multiples_summary.xlsx"

MODEL = "claude-opus-4-8"
MAX_TOKENS = 16000
MAX_SEARCHES_PER_DEAL = 3  # hard cap; the dominant cost is search-result tokens
MAX_PAUSE_CONTINUATIONS = 5

SYSTEM_PROMPT = """\
You are a FinTech M&A research analyst. For the transaction you are given, use
web search to determine whether a valuation multiple was published, or can be
calculated from published information. Work through these steps:

1. Identify the transaction: confirm the actual parties (the names you are
   given may contain typos or have acquirer/target swapped - the search hints
   flag suspected cases, but verify via search), announcement date, deal
   structure (full acquisition, carve-out, take-private, minority/growth
   investment), and whether the deal value was disclosed.
2. Look for a PUBLISHED multiple: an EV/Revenue, EV/ARR, EV/EBITDA, or similar
   multiple stated explicitly in the acquirer's press release, investor
   presentation, SEC/regulatory filing, or reputable press (WSJ, FT, Reuters,
   Bloomberg, trade press like HousingWire or Finextra). Note the exact figure,
   its basis (which metric, which period), and the source.
3. If no multiple was published, determine whether one is CALCULABLE: is the
   deal value (or enterprise value) disclosed, and is the target's revenue,
   ARR, or EBITDA for a relevant period disclosed anywhere public (filings,
   press releases, S-1s, proxy statements, seller's segment reporting, credible
   press reports)? If both inputs exist, compute the implied multiple and show
   the arithmetic. Flag softness in either input (e.g. "revenue reported by
   press, not company-confirmed", "deal value includes earnout").
4. Prefer primary sources. Cite a URL for every figure you rely on. If reports
   conflict, say so and cite both.

SEARCH BUDGET: you have at most <<MAX_SEARCHES>> web searches - plan them, do
not browse. Typically: (1) one query naming both parties to confirm the deal
and its value ("<acquirer> acquires <target> deal value"), (2) one query aimed
at the multiple or the target's financials ("<target> acquisition revenue
multiple" or "<target> ARR revenue"). Spend any remaining search only on one
specific missing figure. Read the result snippets carefully before searching
again - most of what you need is usually in the first set of results. If the
budget runs out before every figure is verified, report what you found, lower
identification_confidence accordingly, and say in "notes" what remains
unverified - never fill gaps by guessing.

Finish your reply with EXACTLY ONE fenced JSON block (```json ... ```) matching
this schema - no other fenced blocks after it:

{
  "deal_id": "<id you were given>",
  "acquirer": "<verified acquirer name>",
  "target": "<verified target name>",
  "announced": "YYYY-MM or null",
  "deal_type": "acquisition | take-private | carve-out | minority/growth | other",
  "identification_confidence": "high | medium | low",
  "deal_value": {
    "disclosed": true/false,
    "value_usd_m": <number or null>,
    "notes": "<currency, earnouts, stake %, or null>"
  },
  "ev_multiple": {
    "published": true/false,
    "value": "<e.g. '18x' or null>",
    "basis": "<e.g. 'EV / 2023E revenue' or null>",
    "source_url": "<url or null>"
  },
  "revenue_multiple": {
    "published": true/false,
    "value": "<e.g. '9.4x' or null>",
    "basis": "<e.g. 'EV / LTM ARR' or null>",
    "source_url": "<url or null>"
  },
  "calculable": {
    "possible": true/false,
    "implied_multiple": "<e.g. '~7.2x EV/Revenue' or null>",
    "calculation": "<arithmetic shown, or null>",
    "inputs": ["<each input figure with its source url>"],
    "caveats": "<softness in the inputs, or null>"
  },
  "verdict": "published | calculable | partially_calculable | not_available",
  "sources": [{"title": "<source>", "url": "<url>"}],
  "notes": "<anything else relevant: conflicting reports, typo resolution, swapped parties>"
}

Verdict rules: "published" if any EV or revenue multiple was stated in a
credible source; "calculable" if not published but both inputs are public and
firm; "partially_calculable" if one input is soft (unconfirmed press figure,
undisclosed earnout, unclear period); "not_available" if deal value or target
financials are simply not public.\
"""

def web_search_tool(max_searches: int) -> dict:
    return {
        "type": "web_search_20260209",
        "name": "web_search",
        "max_uses": max_searches,
    }


def load_deals() -> list[dict]:
    with open(DEALS_FILE, encoding="utf-8") as f:
        return json.load(f)["deals"]


def result_path(deal_id: str) -> Path:
    return RESULTS_DIR / f"{deal_id}.json"


def build_user_prompt(deal: dict) -> str:
    return (
        f"Research this transaction and report on its valuation multiples.\n\n"
        f"deal_id: {deal['id']}\n"
        f"Acquirer (as given): {deal['acquirer']}\n"
        f"Target (as given): {deal['target']}\n"
        f"Source-list spellings: {'; '.join(deal.get('as_given', []))}\n"
        f"Search hints (verify, don't assume): {deal.get('search_hints', 'none')}"
    )


def extract_json_block(text: str) -> dict:
    """Parse the last fenced JSON block out of the model's reply."""
    blocks = re.findall(r"```json\s*(.*?)```", text, re.DOTALL)
    if not blocks:
        # fall back to any fenced block, then to the whole text
        blocks = re.findall(r"```\s*(.*?)```", text, re.DOTALL)
    if not blocks:
        raise ValueError("no fenced JSON block found in model reply")
    return json.loads(blocks[-1])


def research_deal(client, deal: dict, max_searches: int) -> dict:
    """One research pass for one deal, handling pause_turn continuations."""
    messages = [{"role": "user", "content": build_user_prompt(deal)}]
    system = SYSTEM_PROMPT.replace("<<MAX_SEARCHES>>", str(max_searches))

    for _ in range(MAX_PAUSE_CONTINUATIONS + 1):
        response = client.messages.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            thinking={"type": "adaptive"},
            output_config={"effort": "medium"},  # cheaper; raise if quality slips
            system=system,
            tools=[web_search_tool(max_searches)],
            messages=messages,
        )
        if response.stop_reason == "pause_turn":
            # server-side tool loop paused; resend to resume where it left off
            messages = messages[:1] + [
                {"role": "assistant", "content": response.content}
            ]
            continue
        break
    else:
        raise RuntimeError("turn still paused after max continuations")

    if response.stop_reason == "refusal":
        raise RuntimeError("model refused the request")
    if response.stop_reason == "max_tokens":
        raise RuntimeError("response truncated at max_tokens - retry or raise MAX_TOKENS")

    final_text = "\n".join(b.text for b in response.content if b.type == "text")
    result = extract_json_block(final_text)
    result["_usage"] = {
        "input_tokens": response.usage.input_tokens,
        "output_tokens": response.usage.output_tokens,
    }
    return result


def run_research(deals: list[dict], force: bool, limit: int | None, max_searches: int) -> None:
    import anthropic  # deferred so --dry-run / --summarize-only work without the SDK

    client = anthropic.Anthropic(max_retries=4)
    RESULTS_DIR.mkdir(parents=True, exist_ok=True)

    pending = [
        d for d in deals
        if d.get("research", True) and (force or not result_path(d["id"]).exists())
    ]
    if limit:
        pending = pending[:limit]
    if not pending:
        print("Nothing to research - all deals have saved results. Use --force to re-run.")
        return

    print(f"Researching {len(pending)} deal(s) with {MODEL}...")
    failures = []
    for i, deal in enumerate(pending, 1):
        label = f"{deal['acquirer']} / {deal['target']}"
        print(f"[{i}/{len(pending)}] {label} ...", flush=True)
        try:
            result = research_deal(client, deal, max_searches)
        except anthropic.RateLimitError:
            print("  rate limited past SDK retries - stopping; re-run to resume.")
            failures.append(deal["id"])
            break
        except anthropic.APIStatusError as e:
            print(f"  API error {e.status_code}: {e.message}")
            failures.append(deal["id"])
            continue
        except anthropic.APIConnectionError:
            print("  network error - skipping; re-run to retry.")
            failures.append(deal["id"])
            continue
        except (RuntimeError, ValueError, json.JSONDecodeError) as e:
            print(f"  failed: {e}")
            failures.append(deal["id"])
            continue

        with open(result_path(deal["id"]), "w", encoding="utf-8") as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
        print(f"  verdict: {result.get('verdict', '?')}")

    if failures:
        print(f"\n{len(failures)} deal(s) failed: {', '.join(failures)}")
        print("Re-run the script to retry them (saved results are skipped).")


def build_summary(deals: list[dict]) -> None:
    rows = []
    for deal in deals:
        path = result_path(deal["id"])
        if not deal.get("research", True):
            rows.append({
                "deal_id": deal["id"], "acquirer": deal["acquirer"],
                "target": deal["target"], "announced": "", "verdict": "excluded",
                "ev_multiple": "", "revenue_multiple": "", "implied_multiple": "",
                "deal_value_usd_m": "", "notes": deal.get("search_hints", ""),
            })
            continue
        if not path.exists():
            rows.append({
                "deal_id": deal["id"], "acquirer": deal["acquirer"],
                "target": deal["target"], "announced": "", "verdict": "pending",
                "ev_multiple": "", "revenue_multiple": "", "implied_multiple": "",
                "deal_value_usd_m": "", "notes": "",
            })
            continue
        r = json.loads(path.read_text(encoding="utf-8"))
        ev, rev, calc = r.get("ev_multiple", {}), r.get("revenue_multiple", {}), r.get("calculable", {})

        def fmt(m):
            return f"{m.get('value')} ({m.get('basis')})" if m.get("published") else ""

        rows.append({
            "deal_id": r.get("deal_id", deal["id"]),
            "acquirer": r.get("acquirer", deal["acquirer"]),
            "target": r.get("target", deal["target"]),
            "announced": r.get("announced") or "",
            "verdict": r.get("verdict", ""),
            "ev_multiple": fmt(ev),
            "revenue_multiple": fmt(rev),
            "implied_multiple": calc.get("implied_multiple") or "",
            "deal_value_usd_m": (r.get("deal_value") or {}).get("value_usd_m") or "",
            "notes": r.get("notes") or "",
        })

    SUMMARY_CSV.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(rows[0].keys())
    with open(SUMMARY_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    done = [r for r in rows if r["verdict"] not in ("pending", "excluded")]
    lines = [
        "# Precedent Deal Multiples - Research Summary",
        "",
        f"{len(done)} of {len([r for r in rows if r['verdict'] != 'excluded'])} researchable deals completed.",
        "",
        "| Acquirer | Target | Announced | Verdict | EV multiple | Revenue multiple | Implied (calc) | Deal value ($M) |",
        "|---|---|---|---|---|---|---|---|",
    ]
    for r in rows:
        lines.append(
            f"| {r['acquirer']} | {r['target']} | {r['announced']} | {r['verdict']} "
            f"| {r['ev_multiple']} | {r['revenue_multiple']} | {r['implied_multiple']} "
            f"| {r['deal_value_usd_m']} |"
        )
    lines += ["", "Full detail (sources, calculations, caveats) per deal in `outputs/results/*.json`.", ""]
    SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {SUMMARY_CSV.relative_to(HERE)} and {SUMMARY_MD.relative_to(HERE)}")
    build_xlsx(deals)


def build_xlsx(deals: list[dict]) -> None:
    """Excel roll-up with hyperlinked sources: one summary sheet + an all-sources sheet."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed - skipped xlsx export (pip install openpyxl)")
        return

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    link_font = Font(color="0563C1", underline="single")
    wrap = Alignment(wrap_text=True, vertical="top")

    wb = Workbook()
    ws = wb.active
    ws.title = "Deal Multiples"
    headers = [
        "Acquirer", "Target", "Announced", "Deal Type", "Verdict",
        "Deal Value ($M)", "EV Multiple (published)", "Revenue Multiple (published)",
        "Implied Multiple (calculated)", "Calculation", "Caveats", "Source", "Notes",
    ]
    widths = [24, 24, 11, 15, 20, 14, 24, 24, 24, 50, 40, 34, 50]
    ws.append(headers)
    for cell in ws[1]:
        cell.font, cell.fill = header_font, header_fill
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    src_ws = wb.create_sheet("All Sources")
    src_ws.append(["Deal", "Source", "URL"])
    for cell in src_ws[1]:
        cell.font, cell.fill = header_font, header_fill
    for col, w in zip("ABC", [44, 60, 80]):
        src_ws.column_dimensions[col].width = w
    src_ws.freeze_panes = "A2"

    def published_str(m: dict) -> str:
        return f"{m.get('value')} ({m.get('basis')})" if m.get("published") else ""

    for deal in deals:
        path = result_path(deal["id"])
        if not deal.get("research", True) or not path.exists():
            status = "excluded" if not deal.get("research", True) else "pending"
            ws.append([deal["acquirer"], deal["target"], "", "", status,
                       "", "", "", "", "", "", "", deal.get("search_hints", "")])
            continue

        r = json.loads(path.read_text(encoding="utf-8"))
        ev = r.get("ev_multiple") or {}
        rev = r.get("revenue_multiple") or {}
        calc = r.get("calculable") or {}
        sources = r.get("sources") or []
        label = f"{r.get('acquirer', deal['acquirer'])} / {r.get('target', deal['target'])}"

        # primary source: the published multiple's citation, else the first source
        primary_url = (
            (ev.get("source_url") if ev.get("published") else None)
            or (rev.get("source_url") if rev.get("published") else None)
            or (sources[0].get("url") if sources else None)
        )
        primary_title = next(
            (s.get("title") for s in sources if s.get("url") == primary_url and s.get("title")),
            primary_url,
        )

        ws.append([
            r.get("acquirer", deal["acquirer"]),
            r.get("target", deal["target"]),
            r.get("announced") or "",
            r.get("deal_type") or "",
            r.get("verdict") or "",
            (r.get("deal_value") or {}).get("value_usd_m"),
            published_str(ev),
            published_str(rev),
            calc.get("implied_multiple") or "",
            calc.get("calculation") or "",
            calc.get("caveats") or "",
            primary_title or "",
            r.get("notes") or "",
        ])
        row = ws.max_row
        for cell in ws[row]:
            cell.alignment = wrap
        if primary_url:
            src_cell = ws.cell(row=row, column=headers.index("Source") + 1)
            src_cell.hyperlink = primary_url
            src_cell.font = link_font

        for s in sources:
            src_ws.append([label, s.get("title") or "", s.get("url") or ""])
            url_cell = src_ws.cell(row=src_ws.max_row, column=3)
            if s.get("url"):
                url_cell.hyperlink = s["url"]
                url_cell.font = link_font

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    wb.save(SUMMARY_XLSX)
    print(f"Wrote {SUMMARY_XLSX.relative_to(HERE)}")


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument("--deal", help="research a single deal by id")
    parser.add_argument("--limit", type=int, help="cap the number of deals this run")
    parser.add_argument("--force", action="store_true", help="re-run deals that already have results")
    parser.add_argument("--max-searches", type=int, default=MAX_SEARCHES_PER_DEAL,
                        help=f"web-search cap per deal (default {MAX_SEARCHES_PER_DEAL}; raise for a stubborn deal)")
    parser.add_argument("--dry-run", action="store_true", help="show the plan without calling the API")
    parser.add_argument("--summarize-only", action="store_true", help="rebuild the summary from saved results")
    args = parser.parse_args()

    deals = load_deals()
    if args.deal:
        deals = [d for d in deals if d["id"] == args.deal]
        if not deals:
            print(f"Unknown deal id: {args.deal}", file=sys.stderr)
            return 1

    if args.dry_run:
        print(f"{len(deals)} unique deal(s) in deals.json:")
        for d in deals:
            if not d.get("research", True):
                status = "excluded (confidential)"
            elif result_path(d["id"]).exists() and not args.force:
                status = "done"
            else:
                status = "pending"
            dupes = len(d.get("as_given", [])) or 1
            dupe_note = f" [{dupes} source rows]" if dupes > 1 else ""
            print(f"  {status:<24} {d['id']:<28} {d['acquirer']} / {d['target']}{dupe_note}")
        return 0

    if not args.summarize_only:
        run_research(deals, force=args.force, limit=args.limit, max_searches=args.max_searches)

    build_summary(load_deals())
    return 0


if __name__ == "__main__":
    sys.exit(main())
