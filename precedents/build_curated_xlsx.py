#!/usr/bin/env python3
"""Build the curated Deal_Multiples.xlsx: only deals whose multiple is credible.

Inclusion bar: a multiple published by the acquirer or credible press, or a
calculation whose inputs are both firm (deal value from PR/filings + target
financials disclosed by the company/management/seller). Excluded as forced:
Oakley/GLAS (source contradicts its own math), Apax/Finastra TCM (both inputs
unofficial), Fiserv/Finxact (stake-only price / analyst revenue estimate),
Constellation/Optimal Blue (only a 2020-era revenue figure for a 2023 deal),
Inflexion/Infront (equity-only, annualized single quarter), Motive/Backbase
(revenue disclosed only as a floor), plus the 12 deals with no public terms
or target financials. Full detail for every deal: outputs/results/*.json.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
OUT = HERE / "outputs" / "Deal_Multiples.xlsx"

# (target, acquirer, rev multiple, ev multiple, implied (calculated), source title, source url)
ROWS = [
    # -- multiples published --
    ("Adenza", "Nasdaq",
     "~18x 2023E revenue", "~31x 2023E EBITDA", "",
     "CNBC (deal coverage, multiples stated)",
     "https://www.cnbc.com/2023/06/12/nasdaq-to-buy-financial-software-firm-adenza-for-10point5-billion.html"),
    ("OnTheMarket", "CoStar Group",
     "2.5x TTM revenue", "11x TTM adj. EBITDA", "",
     "CoStar Group offer press release",
     "https://www.costargroup.com/press-room/2023/costar-group-offers-acquire-leading-uk-residential-property-portal-onthemarket"),
    ("SitusAMC - CRE Valuation Services unit", "Altus Group",
     "", "13.4x FY2023E EBITDA (net price)", "~4.9x FY2023E revenue (gross price)",
     "Altus Group press release (GlobeNewswire)",
     "https://www.globenewswire.com/news-release/2023/11/09/2777873/0/en/Altus-Group-Enters-into-Agreement-to-Purchase-SitusAMC-s-Commercial-Real-Estate-Valuation-Services-Business.html"),
    ("SimpleNexus", "nCino",
     "28.8x TTM revenue ($1.2B / $41.6M)", "", "",
     "TechBuzz News (revenue per trade press; $1.2B price per nCino PR)",
     "https://www.techbuzznews.com/ncino-acquires-simplenexus-for-1-2-billion/"),
    # -- calculated from firm published inputs --
    ("Mr. Cooper", "Rocket Companies",
     "", "", "~4.2x P/FY2024 revenue; ~14.1x P/E ($9.4B all-stock)",
     "Rocket Companies press release (financials per COOP filings)",
     "https://ir.rocketcompanies.com/news-and-events/press-releases/press-release-details/2025/Mr--Cooper-Americas-Largest-Servicer-Joins-Rocket-the-Nations-Largest-Lender/default.aspx"),
    ("Candescent (NCR Voyix digital banking)", "Veritas Capital",
     "", "", "~4.2x EV/FY2023 segment revenue ($2.45B / ~$579M)",
     "Veritas Capital press release (segment revenue per NCR Voyix disclosures)",
     "https://www.veritascapital.com/ncr-voyix-enters-definitive-agreement-to-sell-digital-banking-to-veritas-capital-for-245-billion-purchase-price/"),
    ("MeridianLink", "Centerbridge",
     "", "", "~6.3x EV/FY2024 revenue; ~15.3x EV/FY2024 adj. EBITDA ($2.0B EV)",
     "MeridianLink press release (financials per MLNK FY2024 results)",
     "https://www.meridianlink.com/press-release/meridianlink-to-be-acquired-by-centerbridge-partners-for-2-0-billion/"),
    ("Technisys", "SoFi",
     "", "", "~15.7x price/CY2021E revenue ($1.1B stock / ~$70M press-reported revenue)",
     "SoFi 8-K press release (SEC; revenue per announcement-day press)",
     "https://www.sec.gov/Archives/edgar/data/0001818874/000181887422000014/exhibit991_8-k2222022.htm"),
    ("Title365", "Blend",
     "", "", "~2.2x equity/FY2020 revenue ($468.5M implied 100% / $212.1M)",
     "Blend Labs S-1 (SEC - both inputs on file)",
     "https://www.sec.gov/Archives/edgar/data/1855747/000119312521194971/d162671ds1.htm"),
    ("Sopra Banking Software", "Axway",
     "", "", "~1.0x EV/FY2023 revenue (EUR 330M / ~EUR 340M)",
     "FinTech Futures (completion coverage)",
     "https://www.fintechfutures.com/m-a/axway-completes-330m-acquisition-of-sopra-banking-software"),
    ("SharpSpring", "Constant Contact (Clearlake/Siris)",
     "", "", "~8.2x EV/FY2020 revenue (~$240M incl. debt / $29.3M)",
     "Constant Contact press release (revenue per SHSP filings)",
     "https://news.constantcontact.com/2021-06-22-Clearlake-Capital-and-Siris-Backed-Constant-Contact-Agrees-to-Acquire-SharpSpring"),
    ("TitlePoint", "Fidelity National Financial",
     "", "", "~5.6x price/annualized Q3'22 revenue ($225M / ~$40M run-rate)",
     "FNF press release (PR Newswire; revenue per Black Knight disclosure)",
     "https://www.prnewswire.com/news-releases/fidelity-national-financial-to-acquire-titlepoint-from-black-knight-301683128.html"),
    ("Blue Water Financial Technologies", "Voxtur Analytics",
     "", "", "~5.6x price/TTM revenue (~8.9x FY2021; $101M consideration)",
     "Voxtur closing press release (GlobeNewswire, financials disclosed)",
     "https://www.globenewswire.com/news-release/2022/09/22/2520882/0/en/Voxtur-Closes-Acquisition-of-Blue-Water-Financial-Technologies-and-Announces-Expansion-of-Credit-Facilities-and-Private-Placement.html"),
    ("Floify", "Porch Group",
     "", "", "~5.8x price/2022E revenue ($86.5M / $15M company estimate)",
     "Porch Group press release (GlobeNewswire)",
     "https://www.globenewswire.com/news-release/2021/10/27/2322103/29193/en/Porch-Group-Acquires-Floify-a-Leading-SaaS-Provider-for-Loan-Officers.html"),
    ("OpenClose", "MeridianLink",
     "", "", "~5.4x price/run-rate revenue ($65M / ~$1M per month per CFO)",
     "National Mortgage News (price + CFO revenue commentary)",
     "https://www.nationalmortgagenews.com/news/meridianlink-buys-openclose-for-65-million-in-bid-to-reach-more-banks"),
    ("Timios", "Ideanomics",
     "", "", "~0.9x price/FY2019 revenue ($40M / $45.1M)",
     "Ideanomics press release (PR Newswire)",
     "https://www.prnewswire.com/news-releases/ideanomics-announces-definitive-agreement-to-acquire-timios-holdings-corp-301171853.html"),
]

HEADERS = ["Target", "Acquirer", "Rev Multiple", "EV Multiple", "Implied", "Source"]
WIDTHS = [38, 34, 32, 30, 62, 52]


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deal Multiples"

    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wrap = Alignment(wrap_text=True, vertical="top")
    link_font = Font(color="0563C1", underline="single")
    for target, acquirer, rev, ev, implied, src_title, src_url in ROWS:
        ws.append([target, acquirer, rev, ev, implied, src_title])
        for cell in ws[ws.max_row]:
            cell.alignment = wrap
        src_cell = ws.cell(row=ws.max_row, column=6)
        src_cell.hyperlink = src_url
        src_cell.font = link_font

    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    wb.save(OUT)
    print(f"Wrote {OUT.relative_to(HERE)} ({len(ROWS)} deals)")


if __name__ == "__main__":
    main()
