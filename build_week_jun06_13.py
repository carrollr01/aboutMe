#!/usr/bin/env python3
"""Build the June 6-13, 2026 deals workbook: Growth Raises + Acquisitions."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RAISE_HEADERS = ["Sector", "Target Country", "Date", "Target", "Lead Investor(s)",
                 "Amount ($M)", "Valuation ($M)", "EV / Revenue", "EV / EBITDA",
                 "Target Description", "Link / Press Release"]
MA_HEADERS = ["Sector", "Target Country", "Deal Type", "Date", "Target", "Acquirer",
              "EV ($M)", "EV / Revenue", "EV / EBITDA", "Target Description",
              "Link / Press Release"]

# Growth raises > $25M, announced 6-13 Jun 2026, core markets.
RAISES = [
 ["Capital Markets Tech","United States","2026-06-11","Digital Asset","a16z crypto",355,2000,"","",
  "A US developer of the Canton Network blockchain for institutional capital markets.",
  "https://www.prnewswire.com/news-releases/digital-asset-raises-355-million-to-accelerate-cantons-role-as-onchain-infrastructure-for-capital-markets-302797427.html"],
 ["Banking & Lending Tech","United States","2026-06-12","Current","Springcoast Partners",80,1500,"","",
  "A US consumer neobank offering banking, spending and credit-building products.",
  "https://www.thebankslate.com/2026/06/neobank-current-raises-80-million-in-latest-funding-round/"],
 ["InsurTech","United States","2026-06-10","Poetic","Kleiner Perkins",50,500,"","",
  "A US developer of deterministic AI for automating complex insurance and enterprise processes.",
  "https://www.prnewswire.com/news-releases/poetic-raises-50m-series-a-to-automate-the-worlds-most-complex-enterprise-processes-with-reliable-ai-302796939.html"],
 ["Payments","United States","2026-06-08","EDGE Markets","CoinFund",29.2,"","","",
  "A US developer of payment-rail and settlement infrastructure for prediction markets and regulated gaming.",
  "https://www.prnewswire.com/news-releases/edge-markets-raises-29-2-million-series-a-funding-round-302793671.html"],
 ["Real Estate & Mortgage Tech","Canada","2026-06-10","nesto","La Caisse",220,1070,"","",
  "A Canadian mortgage technology and financing platform with white-label cloud and AI software for lenders.",
  "https://www.globenewswire.com/news-release/2026/06/10/3309762/0/en/nesto-raises-302-million-Series-E-at-1-47-billion-valuation-to-accelerate-growth.html"],
 ["Financial Info & Analytics","United States","2026-06-11","Hypha","TriEdge Investments",50,"","","",
  "A US AI-native investment-intelligence platform structuring data across private credit and private equity.",
  "https://www.businesswire.com/news/home/20260611628926/en/Hypha-Emerges-From-Stealth-Announces-a-$50M-Seed-Round"],
 ["Banking & Lending Tech","Canada","2026-06-11","KOHO","Mubadala",95,970,"","",
  "A Canadian consumer neobank offering spending, savings and credit-building products.",
  "https://betakit.com/koho-becomes-canadas-latest-unicorn-following-130-million-series-e-round/"],
]

# Acquisitions (any size), announced 6-13 Jun 2026, core markets.
ACQUISITIONS = [
 ["Asset & Wealth Tech","Switzerland","Strategic M&A","2026-06-08","additiv","Temenos","","","",
  "A Swiss orchestration platform for wealth management and financial services.",
  "https://www.temenos.com/press_release/temenos-acquires-additiv-to-strengthen-its-wealth-proposition-and-accelerate-ai-driven-orchestration/"],
 ["Capital Markets Tech","United States","Strategic M&A","2026-06-08","LevPro","Octus","","","",
  "A US front-office software provider for CLO and private-credit portfolio management and trading.",
  "https://octus.com/company/news/octus-signs-definitive-agreement-to-acquire-levpro/"],
 ["Real Estate & Mortgage Tech","United States","Strategic M&A","2026-06-10","Kiavi","Figure",717,"","",
  "A US AI-powered lending platform for residential real-estate investors.",
  "https://www.kiavi.com/press/figure-acquires-kiavi"],
 ["Payments","United States","Strategic M&A","2026-06-12","Payoneer","Nuvei",2750,"","",
  "A US-based global cross-border payments and money-movement platform for SMBs.",
  "https://www.prnewswire.com/news-releases/nuvei-to-acquire-payoneer-for-2-75-billion-creating-a-leading-global-platform-for-local-and-cross-border-commerce-302800166.html"],
 ["Corporate Financial Function","United States","Strategic M&A","2026-06-11","Orb","Adyen",335,"","",
  "A US enterprise billing and revenue-management platform for usage-based pricing.",
  "https://www.adyen.com/press-and-media/jtrg4qd7j3p4rj"],
 ["Corporate Financial Function","United Kingdom","Strategic M&A","2026-06-08","m3ter","Salesforce","","","",
  "A UK usage-based billing and revenue-metering platform for software companies.",
  "https://www.salesforce.com/news/stories/salesforce-signs-definitive-agreement-to-acquire-m3ter/"],
 ["Financial Info & Analytics","United States","Strategic M&A","2026-06-11","Messari","Blockworks",10,"","",
  "A US crypto market-data, research and intelligence platform.",
  "https://www.businesswire.com/news/home/20260611031601/en/Blockworks-Acquires-Messari-Combining-the-Two-Largest-Crypto-Data-Platforms"],
 ["Financial Info & Analytics","United States","Strategic M&A","2026-06-10","RiskFront AI","K2 Integrity","","","",
  "A US developer of agentic AI for financial-crime (AML) compliance and risk operations.",
  "https://fintech.global/2026/06/10/k2-integrity-targets-financial-crime-with-ai-deal/"],
]

header_fill = PatternFill("solid", fgColor="1F6FB2")
header_font = Font(color="FFFFFF", bold=True, size=11)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill_sheet(ws, headers, rows, widths):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = center; cell.border = border
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=1 + len(rows), min_col=1, max_col=len(headers)):
        for cell in row:
            cell.alignment = left_wrap; cell.border = border
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{1 + len(rows)}"
    ws.row_dimensions[1].height = 30


wb = Workbook()
fill_sheet(wb.active, RAISE_HEADERS, RAISES, [22, 16, 12, 24, 24, 11, 14, 12, 12, 52, 46])
wb.active.title = "Growth Raises"
fill_sheet(wb.create_sheet("Acquisitions"), MA_HEADERS, ACQUISITIONS, [24, 16, 18, 12, 22, 16, 9, 12, 12, 52, 46])
wb.save("/home/user/aboutMe/fintech-deals-2026-06-06_06-13.xlsx")

with open("/home/user/aboutMe/fintech-growth-raises-2026-06-06_06-13.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(RAISE_HEADERS); w.writerows(RAISES)
with open("/home/user/aboutMe/fintech-acquisitions-2026-06-06_06-13.csv", "w", newline="") as f:
    w = csv.writer(f); w.writerow(MA_HEADERS); w.writerows(ACQUISITIONS)

print(f"Wrote week 06-06..06-13: {len(RAISES)} raises + {len(ACQUISITIONS)} acquisitions")
