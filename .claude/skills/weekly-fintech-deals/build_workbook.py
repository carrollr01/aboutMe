#!/usr/bin/env python3
"""
build_workbook.py — Weekly FinTech deals workbook, matched to the finalized-DB schema.

Two tabs, exact names and column order:
  "All M&A (PE & Strategic)":
     x | Week | Sector | Target Country | Deal Type | Date | Target | Acquirer | EV ($M) |
     EV / Revenue | EV / EBITDA | Target Description | Link / Press Release |
     Mults Source | Mults Basis | Public Deal | HL Deal | Seller
  "Growth Capital Raises":
     x | Week | Sector | Target Country | Date | Target | Lead Investor(s) | Amount ($M) |
     Valuation ($M) | EV / Revenue | EV / EBITDA | Target Description | Link / Press Release

Conventions (locked):
  - Deal Type is STRICTLY "Strategic M&A" or "PE Buyout" (validated; build fails otherwise).
  - Undisclosed / not-applicable = "-" (never blank) in EV/Amount/Valuation/multiples and
    the M&A meta columns.
  - All money in USD $M (convert at announcement-date FX BEFORE building; see fx note).
  - Week = the deal's Saturday-Friday week, labeled by ending Friday (DD-Mmm-YY), derived
    from Date automatically.
  - Meta columns (Mults Source/Mults Basis/Public Deal/HL Deal/Seller): populate what is
    publicly knowable; "-" otherwise.
  - Link column renders the word "Link" hyperlinked to the source.

Usage: python3 build_workbook.py deals.json "Weekly_Fintech_Deals_<endingFriday>.xlsx"

JSON schema:
{
  "ma": [{ "sector","country","deal_type","date"("DD-Mmm-YY"),"target","acquirer",
           "ev","ev_rev","ev_ebitda","description","link",
           "mults_source","mults_basis","public_deal","hl_deal","seller" }],
  "raises": [{ "sector","country","date","target","lead","amount","valuation",
               "ev_rev","ev_ebitda","description","link" }]
}
Any numeric field may be a number or "-". Missing keys default to "-".
"""
import json, sys, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ALLOWED_DEAL_TYPES = {"Strategic M&A", "PE Buyout"}

MA_HEADERS = ["x","Week","Sector","Target Country","Deal Type","Date","Target","Acquirer",
              "EV ($M)","EV / Revenue","EV / EBITDA","Target Description","Link / Press Release",
              "Mults Source","Mults Basis","Public Deal","HL Deal","Seller"]
MA_KEYS    = ["x","week","sector","country","deal_type","date","target","acquirer",
              "ev","ev_rev","ev_ebitda","description","link",
              "mults_source","mults_basis","public_deal","hl_deal","seller"]

RA_HEADERS = ["x","Week","Sector","Target Country","Date","Target","Lead Investor(s)",
              "Amount ($M)","Valuation ($M)","EV / Revenue","EV / EBITDA",
              "Target Description","Link / Press Release"]
RA_KEYS    = ["x","week","sector","country","date","target","lead",
              "amount","valuation","ev_rev","ev_ebitda","description","link"]

DASH_KEYS = {"ev","ev_rev","ev_ebitda","amount","valuation",
             "mults_source","mults_basis","public_deal","hl_deal","seller"}

HF = PatternFill("solid", fgColor="1F6FB2"); HFONT = Font(color="FFFFFF", bold=True, size=11)
LF = Font(color="0563C1", underline="single", size=11)
CT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LW = Alignment(horizontal="left", vertical="top", wrap_text=True)
LC = Alignment(horizontal="center", vertical="top")
TH = Side(style="thin", color="D0D0D0"); BD = Border(TH, TH, TH, TH)


def week_ending_friday(date_str):
    """DD-Mmm-YY -> that week's ending Friday as DD-Mmm-YY."""
    try:
        d = datetime.datetime.strptime(str(date_str).strip(), "%d-%b-%y").date()
    except ValueError:
        return "-"
    fri = d + datetime.timedelta(days=(4 - d.weekday()) % 7)
    return fri.strftime("%d-%b-%y")


def norm(row, key):
    v = row.get(key, None)
    if key == "week" and (v in (None, "")):
        return week_ending_friday(row.get("date", ""))
    if key == "x":
        return v if v not in (None, "") else ""
    if v in (None, ""):
        return "-" if key in DASH_KEYS else ""
    return v


def fill_sheet(ws, headers, keys, rows, widths):
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        x = ws.cell(1, c); x.fill = HF; x.font = HFONT; x.alignment = CT; x.border = BD
    link_idx = keys.index("link")
    for ri, row in enumerate(rows, 2):
        if "deal_type" in keys:
            dt = row.get("deal_type", "")
            if dt not in ALLOWED_DEAL_TYPES:
                raise ValueError(f"Deal Type must be one of {ALLOWED_DEAL_TYPES}, got: {dt!r} ({row.get('target')})")
        for ci, key in enumerate(keys, 1):
            x = ws.cell(ri, ci); x.border = BD
            if ci - 1 == link_idx:
                url = row.get("link", "")
                x.value = "Link" if url else "-"
                if url:
                    x.hyperlink = url; x.font = LF
                x.alignment = LC
            else:
                v = norm(row, key)
                x.value = v; x.alignment = LW
                if key in ("ev_rev", "ev_ebitda") and isinstance(v, (int, float)):
                    x.number_format = '0.0"x"'
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{1 + len(rows)}"


def sort_key(row):
    """Sector alphabetical, then most recent date first within the sector."""
    try:
        d = datetime.datetime.strptime(str(row.get("date", "")).strip(), "%d-%b-%y").date()
    except ValueError:
        d = datetime.date.min
    return (row.get("sector", ""), -d.toordinal())


def main():
    data_path = sys.argv[1] if len(sys.argv) > 1 else "deals.json"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "Weekly_Fintech_Deals.xlsx"
    with open(data_path) as f:
        data = json.load(f)
    ma = sorted(data.get("ma", []), key=sort_key)
    ra = sorted(data.get("raises", []), key=sort_key)
    wb = Workbook()
    ws = wb.active; ws.title = "All M&A (PE & Strategic)"
    fill_sheet(ws, MA_HEADERS, MA_KEYS, ma,
               [4, 10, 22, 14, 14, 10, 18, 18, 9, 11, 11, 46, 9, 14, 14, 10, 8, 16])
    ws2 = wb.create_sheet("Growth Capital Raises")
    fill_sheet(ws2, RA_HEADERS, RA_KEYS, ra,
               [4, 10, 22, 14, 10, 18, 26, 11, 13, 11, 11, 46, 9])
    wb.save(out_path)
    print(f"Wrote {out_path}: {len(ma)} M&A + {len(ra)} raises")


if __name__ == "__main__":
    main()
